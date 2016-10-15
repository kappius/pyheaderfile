#!/usr/bin/env python
# -*- coding: utf-8 -*-

from .basefile import PyHeaderSheet
from .utils import is_str_or_unicode

class Xls(PyHeaderSheet):
    """
        class that read xls files

        >>> test = Xls(name="test", header=["col1","col2","col3"])
        >>> test.write(*["test1","test2","test3"])
        >>> test.save('../')
        >>> test = Xls(name='../test.xls')
        >>> content = test.read()
        >>> sorted(next(content).items())
        [('col1', 'test1'), ('col2', 'test2'), ('col3', 'test3')]
        >>> test.name = 'test2'
        >>> convert_xlsx = Xlsx()
        >>> convert_xlsx(test)
        >>> convert_xlsx.save()
        >>> from .headercsv import Csv
        >>> convert_csv = Csv()
        >>> convert_csv(test)
        >>> convert_csv.save()

    """
    def __init__(self, name=None, header=list(), sheet_name=None, style=None,
                 strip=False):
        self.name = name
        self.header = header
        self.sheet_name = sheet_name
        self.style = style
        self.strip = strip
        self.colors = dict()
        super(Xls, self).__init__()

    def read_cell(self, x, y):
        """
            reads the cell at position x and y; puts the default styles in xlwt
        """
        cell = self._sheet.row(x)[y]
        if self._file.xf_list[
            cell.xf_index].background.pattern_colour_index == 64:
            self._file.xf_list[
                cell.xf_index].background.pattern_colour_index = 9
        if self._file.xf_list[
            cell.xf_index].background.pattern_colour_index in self.colors.keys():
            style = self.colors[self._file.xf_list[
                cell.xf_index].background.pattern_colour_index]
        else:
            style = self.xlwt.easyxf(
                'pattern: pattern solid; border: top thin, right thin, bottom thin, left thin;')
            style.pattern.pattern_fore_colour = self._file.xf_list[
                cell.xf_index].background.pattern_colour_index
            self.colors[self._file.xf_list[
                cell.xf_index].background.pattern_colour_index] = style
        style.font.name = self._file.font_list[
            self._file.xf_list[cell.xf_index].font_index].name
        style.font.bold = self._file.font_list[
            self._file.xf_list[cell.xf_index].font_index].bold
        if isinstance(self.header[y], tuple):
            header = self.header[y][0]
        else:
            header = self.header[y]
        if self.strip:
            if is_str_or_unicode(cell.value):
                cell.value = cell.value.strip()
        if self.style:
            return {header: (cell.value, style)}
        else:
            return {header: cell.value}

    def write_cell(self, x, y, value, style=None):
        """
            writing style and value in the cell of x and y position
        """
        if isinstance(style, str):
            style = self.xlwt.easyxf(style)
        if style:
            self._sheet.write(x, y, label=value, style=style)
        else:
            self._sheet.write(x, y, label=value)

    def close(self):
        # save and close without changing path
        self._file.save(self.name)
        if not is_str_or_unicode(self.name):
            return self.name.getvalue()

    def save(self, path=None):
        # save the file
        if is_str_or_unicode(self.name):
            name = self.name
        else:
            name = 'default.xls'

        if path:
            basename = self.path.basename(name)
            self._file.save(self.path.join(path, basename))
        else:
            return self.close()

    def _create(self):
        # create the file and sheet; write the header
        self._file = self.xlwt.Workbook(style_compression=2)
        if is_str_or_unicode(self.name):
            name = self.path.splitext(self.name)[0]
            basename = self.path.basename(name)
            if not self.sheet_name:
                self.sheet_name = basename
            self.name = "%s.xls" % name
        else:
            self.sheet_name = self.sheet_name or 'default'

        self._sheet = self._file.add_sheet(sheetname=self.sheet_name,
                                           cell_overwrite_ok=True)
        self.write(*self.header)

    def _open(self):
        # open the file and get sheets
        if not hasattr(self, '_file'):
            if is_str_or_unicode(self.name):
                self._file = self.xlrd.open_workbook(filename=self.name,
                                                     formatting_info=True)
            else:
                self._file = self.xlrd.open_workbook(file_contents=self.name.getvalue(),
                                                     formatting_info=True)
            self.sheet_names = self._file.sheet_names()

    def _open_sheet(self):
        # read the sheet, get value the header, get number columns and rows
        if self.sheet_name and not self.header:
            self._sheet = self._file.sheet_by_name(self.sheet_name)
            self.header = [cell.value for cell in self._sheet.row(0)]
            self.ncols = self._sheet.ncols
            self.nrows = self._sheet.nrows

    def _import(self):
        import xlrd
        import xlwt
        import os.path

        self.path = os.path
        self.xlrd = xlrd
        self.xlwt = xlwt


class Xlsx(PyHeaderSheet):
    """
        class that read xlsx files

        >>> test = Xlsx(name="test", header=["col1","col2","col3"])
        >>> test.write(*["test1","test2","test3"])
        >>> test.save('../')
        >>> test = Xlsx(name='../test.xlsx')
        >>> content = test.read()
        >>> sorted(next(content).items())
        [('col1', 'test1'), ('col2', 'test2'), ('col3', 'test3')] 
        >>> test.name = 'test2'
        >>> convert_xls = Xls()
        >>> convert_xls(test)
        >>> convert_xls.save()
        >>> from .headercsv import Csv
        >>> convert_csv = Csv()
        >>> convert_csv(test)
        >>> convert_csv.save()

    """

    def __init__(self, name=None, header=list(), sheet_name=None, style=None,
                 strip=False):
        self.name = name
        self.header = header
        self.style = style
        self.strip = strip
        self.sheet_name = sheet_name
        super(Xlsx, self).__init__()

    def read_cell(self, x, y):
        # reads the cell at position x and y; return value and style
        if isinstance(self.header[y], tuple):
            header = self.header[y][0]
        else:
            header = self.header[y]
        if self.strip:
            if is_str_or_unicode(list(self._sheet.rows)[x][y].value):
                self._sheet.rows[x][y].value = list(self._sheet.rows)[x][y].value.strip()
        if self.style:
            return {header: (
                self._sheet.rows[x][y].value, list(self._sheet.rows)[x][y].style)}
        else:
            return {header: list(self._sheet.rows)[x][y].value}

    def write_cell(self, x, y, value, style=None):
        # writing style and value in the cell of x+1 and y+1 position
        self._sheet.cell(row=x + 1, column=y + 1).value = value
        if style:
            self._sheet.cell(row=x + 1, column=y + 1).style = style

    def close(self):
        # save and close without changing path
        self._file.save(self.name)
        if not is_str_or_unicode(self.name):
            return self.name.getvalue()

    def save(self, path=None):
        # save the file
        if is_str_or_unicode(self.name):
            name = self.name
        else:
            name = 'default.xlsx'

        if path:
            basename = self.path.basename(name)
            self._file.save(self.path.join(path,basename))
        else:
            return self.close()

    def _open(self):
        # open the file with the function xlwt and openpyxl; get sheets
        if not hasattr(self, '_file'):
            #  needed to get right col number
            if is_str_or_unicode(self.name):
                self.file_xlrd = self.xlrd.open_workbook(filename=self.name,
                                                     formatting_info=False)
            else:
                self.file_xlrd = self.xlrd.open_workbook(file_contents=self.name.getvalue(),
                                                     formatting_info=False)

            self._file = self.openpyxl.load_workbook(filename=self.name)
            self.sheet_names = self._file.get_sheet_names()

    def _open_sheet(self):
        # read the sheet, get value the header, get number columns and rows
        if self.sheet_name and not self.header:
            self._sheet = self._file.get_sheet_by_name(self.sheet_name)
            self.sheet_xlrd = self.file_xlrd.sheet_by_name(self.sheet_name)
            self.ncols = self.sheet_xlrd.ncols
            self.nrows = self.sheet_xlrd.nrows
            for i in range(0, self.ncols):
                self.header = self.header + [list(self._sheet.rows)[0][i].value]

    def _create(self):
        # create the file and sheet; write the header
        self._file = self.openpyxl.Workbook()
        if is_str_or_unicode(self.name):
            name = self.path.splitext(self.name)[0]
            basename = self.path.basename(name)
            if not self.sheet_name:
                self.sheet_name = basename
            self.name = "%s.xlsx" % name
        else:
            self.sheet_name = self.sheet_name or 'default'

        self._sheet = self._file.active
        self._sheet.title = self.sheet_name
        self.write(*self.header)

    def _import(self):
        import openpyxl
        import xlrd
        import os.path

        self.path = os.path
        self.openpyxl = openpyxl
        self.xlrd = xlrd
