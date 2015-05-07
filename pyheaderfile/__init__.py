#!/usr/bin/env python
# -*- coding: utf-8 -*-

__all__ = ['Csv', 'Xls', 'Xlsx', 'guess_type']

VERSION = (0, 2, 6)
__version__ = ".".join(map(str, VERSION))


class PyHeaderFile(object):
    # father class of all filetypes

    def __init__(self, *args, **kwargs):
        self._import()
        if self.name:
            if self.header:
                self._create()
            self._open()

    def __call__(self, instance, **kwargs):

        # convert any File object to any File and save it

        self.__init__(instance.name, instance.header, **kwargs)

        for line in instance.read():
            if isinstance(line[instance.header[0]], tuple):
                new_line = dict()
                for l in line:
                    new_line[l] = line[l][0]
                self.write(**new_line)
            else:
                self.write(**line)
        self.save()

    def read(self):

        # read each line of file. Should be an interator

        raise NotImplementedError

    def write(self, *args, **kwargs):
        """
        write to file by args or kwargs. Should be a list
        with elements or a dict for writes with unordered
        lines
        """
        raise NotImplementedError

    def __exit__(self):
        self.save()

    def save(self, path=None):

        # save and close file in other path

        return NotImplemented

    # can use method close or save
    def close(self):

        # save and close file

        return NotImplemented


    # getter and setter for filename. You can change filename to convert
    @property
    def name(self):
        return self._name

    @name.setter
    def name(self, name):
        self._name = name

    # getter and setter for file header
    @property
    def header(self):
        return self._header

    @header.setter
    def header(self, header):
        self._header = header

    def _open(self):
        """
        open file and get header of file. Some oder needed initialization
        things can be done here
        """
        raise NotImplementedError

    def _create(self):

        # create new file with right extension

        raise NotImplementedError

    def _import(self):

        # import needed libs. This prevent conflicts

        raise NotImplementedError


class Csv(PyHeaderFile):
    """
        class that read csv files with ; and , and #

        >>> type(Csv())
        <class '__main__.Csv'>
        >>> test = Csv(name="test", header=["col1","col2","col3"])
        >>> test.write(*["test1","test2","test3"])
        >>> test.save('../')
        >>> test = Csv(name='../test.csv')
        >>> [i for i in test.read()]
        [{u'col2': u'test2', u'col3': u'test3', u'col1': u'test1'}]
        >>> test.name = 'test2'
        >>> convert_xlsx = Xlsx()
        >>> convert_xlsx(test)
        >>> convert_xlsx.save()
        >>> convert_xls = Xls()
        >>> convert_xls(test)
        >>> convert_xls.save()

    """

    def __init__(self, name=None, header=list(), encode='utf-8', header_line=0,
                 delimiters=[",", ";", "#"], strip=False,
                 quotechar='"'):
        self.name = name
        self.header = header
        self.delimiters = list(delimiters)
        self.quotechar = quotechar
        self.encode = encode
        self.header_line = header_line
        self.strip = strip
        super(Csv, self).__init__()

    def read(self):
        # open file in mode write and read the line. Return the dict 'header = value'
        if isinstance(self.name, str) or isinstance(self.name, unicode):
            if not hasattr(self, '_file'):
                self._open()
            elif self._file.mode == 'wb':
                self._file.close()
                self._open()
        else:
            self._open()
        for row in self.reader:
            if self.strip:
                row = [r.strip() for r in row]
            yield dict(zip(self.header, row))

    def write(self, *args, **kwargs):
        # write the value in the file
        if isinstance(self.name, str) or isinstance(self.name, unicode):
            if not hasattr(self, '_file'):
                self._file = open(self.name, 'a')
            elif self._file.mode == 'rb':
                self._file.close()
                self._file = open(self.name, 'a')
        else:
            self._file = self.name
        writer = self.csv.DictWriter(self._file, delimiter=self.delimiters[0],
                                     fieldnames=self.header,
                                     quotechar=self.quotechar,
                                     quoting=self.csv.QUOTE_MINIMAL)
        if args:
            kwargs = dict(zip(self.header, args))
        writer.writerow(kwargs)

    def close(self):
        # close de file
        if not (isinstance(self.name, str) or isinstance(self.name, unicode)):
            return self.name.getvalue()

        if hasattr(self, '_file'):
            self._file.close()

    def save(self, path=None):
        # move and close de file
        if (isinstance(self.name, str) or isinstance(self.name, unicode)) and path:
            self.close()
            name = self.name
            basename = self.path.basename(name)
            self.rename(name, self.path.join(path, basename))
        elif path:
            name = 'default.csv'
            content = self.close()
            with open(self.path.join(path, name), 'w') as f:
                f.write(content)
        else:
           return self.close()

    def _get_dialect(self):
        # discover a dialect to csv file based on some delimiters
        try:
            for i in range(0, self.header_line):
                self._file.readline()
            self.dialect = self.csv.Sniffer().sniff(self._file.readline(),
                                                    delimiters=self.delimiters)
        except:
            self.dialect = self.delimiters[0]
        self._file.seek(0)

    def _open(self):
        # open the file and get header
        if isinstance(self.name, str) or isinstance(self.name, unicode):
            self._file = open(self.name, 'rb')
        else:
            self._file = self.name
        self._file.seek(0)
        self._get_dialect()
        self.reader = self.csv.reader(self._file, self.dialect,
                                      encoding=self.encode, doublequote=True)
        for i in range(0, self.header_line):
            self.reader.next()
        self.header = self.reader.next()

    def _create(self):
        # create the file and write the header
        if isinstance(self.name, str) or isinstance(self.name, unicode):
            name = self.path.splitext(self.name)[0]
            self.name = "%s.csv" % name
            self._file = open(self.name, 'wb')
        else:
            self._file = self.name

        if isinstance(self.name, str) or isinstance(self.name, unicode):
            self._file.seek(0)
            self.write(*self.header)
            self._file.close()
        else:
            self.write(*self.header)

    def _import(self):
        import unicodecsv as csv
        import os

        self.rename = os.rename
        self.csv = csv
        self.path = os.path


class PyHeaderSheet(PyHeaderFile):
    # class that use similar functions for sheets
    def __init__(self):
        self._row = 0
        super(PyHeaderSheet, self).__init__()
        if self.header:
            if isinstance(self.header[0], tuple):
                self.header = [h[0] for h in self.header]
        if not self.sheet_name and self.name:
            self._first_sheet()
        self._open_sheet()

    # define and get sheet name into a spreadsheet file
    @property
    def sheet_name(self):
        return self._sheet_name

    @sheet_name.setter
    def sheet_name(self, sheet_name):
        self._sheet_name = sheet_name

    def _first_sheet(self):
        # get first sheet
        try:
            self.sheet_name = self.sheet_names[0]
        except IndexError:
            raise Exception('There are no sheets')

    @property
    def sheet_names(self):
        # returns a list with the sheet file
        return self._sheet_names

    @sheet_names.setter
    def sheet_names(self, sheet_names):
        # returns a list with the sheet file
        self._sheet_names = sheet_names

    def read(self):
        # read the file line
        for x in xrange(1, self.nrows):
            row = dict()
            for y in xrange(0, self.ncols):
                row.update(self.read_cell(x, y))
            yield row

    # pass x, y, value and style for function write_cell
    def write(self, *args, **kwargs):
        """
        :param args: tuple(value, style), tuple(value, style)
        :param kwargs: header=tuple(value, style), header=tuple(value, style)
        :param args: value, value
        :param kwargs: header=value, header=value
        """

        if args:
            kwargs = dict(zip(self.header, args))
        for header in kwargs:
            cell = kwargs[header]
            if not isinstance(cell, tuple):
                cell = (cell,)
            self.write_cell(self._row, self.header.index(header), *cell)
        self._row += 1


class Xls(PyHeaderSheet):
    """
        class that read xls files

        >>> type(Xls())
        <class '__main__.Xls'>
        >>> test = Xls(name="test", header=["col1","col2","col3"])
        >>> test.write(*["test1","test2","test3"])
        >>> test.save('../')
        >>> test = Xls(name='../test.xls')
        >>> [i for i in test.read()]
        [{u'col2': u'test2', u'col3': u'test3', u'col1': u'test1'}]
        >>> test.name = 'test2'
        >>> convert_xlsx = Xlsx()
        >>> convert_xlsx(test)
        >>> convert_xlsx.save()
        >>> convert_csv = Csv()
        >>> convert_csv(test)
        >>> convert_csv.save()

    """
    def __init__(self, name=None, header=list(), sheet_name=None, style=None,
                 strip=False):
        self.name = name
        self.header = header
        self.sheet_name = sheet_name
        self.style = style
        self.strip = strip
        self.colors = dict()
        super(Xls, self).__init__()

    def read_cell(self, x, y):
        """
            reads the cell at position x and y; puts the default styles in xlwt
        """
        cell = self._sheet.row(x)[y]
        if self._file.xf_list[
            cell.xf_index].background.pattern_colour_index == 64:
            self._file.xf_list[
                cell.xf_index].background.pattern_colour_index = 9
        if self._file.xf_list[
            cell.xf_index].background.pattern_colour_index in self.colors.keys():
            style = self.colors[self._file.xf_list[
                cell.xf_index].background.pattern_colour_index]
        else:
            style = self.xlwt.easyxf(
                'pattern: pattern solid; border: top thin, right thin, bottom thin, left thin;')
            style.pattern.pattern_fore_colour = self._file.xf_list[
                cell.xf_index].background.pattern_colour_index
            self.colors[self._file.xf_list[
                cell.xf_index].background.pattern_colour_index] = style
        style.font.name = self._file.font_list[
            self._file.xf_list[cell.xf_index].font_index].name
        style.font.bold = self._file.font_list[
            self._file.xf_list[cell.xf_index].font_index].bold
        if isinstance(self.header[y], tuple):
            header = self.header[y][0]
        else:
            header = self.header[y]
        if self.strip:
            if is_str_or_unicode(cell.value):
                cell.value = cell.value.strip()
        if self.style:
            return {header: (cell.value, style)}
        else:
            return {header: cell.value}

    def write_cell(self, x, y, value, style=None):
        """
            writing style and value in the cell of x and y position
        """
        if isinstance(style, str):
            style = self.xlwt.easyxf(style)
        if style:
            self._sheet.write(x, y, label=value, style=style)
        else:
            self._sheet.write(x, y, label=value)

    def close(self):
        # save and close without changing path
        self._file.save(self.name)
        if not is_str_or_unicode(self.name):
            return self.name.getvalue()

    def save(self, path=None):
        # save the file
        if is_str_or_unicode(self.name):
            name = self.name
        else:
            name = 'default.xls'

        if path:
            basename = self.path.basename(name)
            self._file.save(self.path.join(path, basename))
        else:
            return self.close()

    def _create(self):
        # create the file and sheet; write the header
        self._file = self.xlwt.Workbook(style_compression=2)
        if is_str_or_unicode(self.name):
            name = self.path.splitext(self.name)[0]
            basename = self.path.basename(name)
            if not self.sheet_name:
                self.sheet_name = basename
            self.name = "%s.xls" % name
        else:
            self.sheet_name = self.sheet_name or 'default'

        self._sheet = self._file.add_sheet(sheetname=self.sheet_name,
                                           cell_overwrite_ok=True)
        self.write(*self.header)

    def _open(self):
        # open the file and get sheets
        if not hasattr(self, '_file'):
            if is_str_or_unicode(self.name):
                self._file = self.xlrd.open_workbook(filename=self.name,
                                                     formatting_info=True)
            else:
                self._file = self.xlrd.open_workbook(file_contents=self.name.getvalue(),
                                                     formatting_info=True)
            self.sheet_names = self._file.sheet_names()

    def _open_sheet(self):
        # read the sheet, get value the header, get number columns and rows
        if self.sheet_name and not self.header:
            self._sheet = self._file.sheet_by_name(self.sheet_name)
            self.header = [cell.value for cell in self._sheet.row(0)]
            self.ncols = self._sheet.ncols
            self.nrows = self._sheet.nrows

    def _import(self):
        import xlrd
        import xlwt
        import os.path

        self.path = os.path
        self.xlrd = xlrd
        self.xlwt = xlwt


class Xlsx(PyHeaderSheet):
    """
        class that read xlsx files

        >>> type(Xlsx())
        <class '__main__.Xlsx'>
        >>> test = Xlsx(name="test", header=["col1","col2","col3"])
        >>> test.write(*["test1","test2","test3"])
        >>> test.save('../')
        >>> test = Xlsx(name='../test.xlsx')
        >>> [i for i in test.read()]
        [{u'col2': u'test2', u'col3': u'test3', u'col1': u'test1'}]
        >>> test.name = 'test2'
        >>> convert_xls = Xls()
        >>> convert_xls(test)
        >>> convert_xls.save()
        >>> convert_csv = Csv()
        >>> convert_csv(test)
        >>> convert_csv.save()

    """

    def __init__(self, name=None, header=list(), sheet_name=None, style=None,
                 strip=False):
        self.name = name
        self.header = header
        self.style = style
        self.strip = strip
        self.sheet_name = sheet_name
        super(Xlsx, self).__init__()

    def read_cell(self, x, y):
        # reads the cell at position x and y; return value and style
        if isinstance(self.header[y], tuple):
            header = self.header[y][0]
        else:
            header = self.header[y]
        if self.strip:
            if is_str_or_unicode(self._sheet.rows[x][y].value):
                self._sheet.rows[x][y].value = self._sheet.rows[x][y].value.strip()
        if self.style:
            return {header: (
                self._sheet.rows[x][y].value, self._sheet.rows[x][y].style)}
        else:
            return {header: self._sheet.rows[x][y].value}

    def write_cell(self, x, y, value, style=None):
        # writing style and value in the cell of x+1 and y+1 position
        self._sheet.cell(row=x + 1, column=y + 1).value = value
        if style:
            self._sheet.cell(row=x + 1, column=y + 1).style = style

    def close(self):
        # save and close without changing path
        self._file.save(self.name)
        if not is_str_or_unicode(self.name):
            return self.name.getvalue()

    def save(self, path=None):
        # save the file
        if is_str_or_unicode(self.name):
            name = self.name
        else:
            name = 'default.xlsx'

        if path:
            basename = self.path.basename(name)
            self._file.save(self.path.join(path,basename))
        else:
            return self.close()

    def _open(self):
        # open the file with the function xlwt and openpyxl; get sheets
        if not hasattr(self, '_file'):
            #  needed to get right col number
            if is_str_or_unicode(self.name):
                self.file_xlrd = self.xlrd.open_workbook(filename=self.name,
                                                     formatting_info=False)
            else:
                self.file_xlrd = self.xlrd.open_workbook(file_contents=self.name.getvalue(),
                                                     formatting_info=False)

            self._file = self.openpyxl.load_workbook(filename=self.name)
            self.sheet_names = self._file.get_sheet_names()

    def _open_sheet(self):
        # read the sheet, get value the header, get number columns and rows
        if self.sheet_name and not self.header:
            self._sheet = self._file.get_sheet_by_name(self.sheet_name)
            self.sheet_xlrd = self.file_xlrd.sheet_by_name(self.sheet_name)
            self.ncols = self.sheet_xlrd.ncols
            self.nrows = self.sheet_xlrd.nrows
            for i in range(0, self.ncols):
                self.header = self.header + [self._sheet.rows[0][i].value]

    def _create(self):
        # create the file and sheet; write the header
        self._file = self.openpyxl.Workbook()
        if is_str_or_unicode(self.name):
            name = self.path.splitext(self.name)[0]
            basename = self.path.basename(name)
            if not self.sheet_name:
                self.sheet_name = basename
            self.name = "%s.xlsx" % name
        else:
            self.sheet_name = self.sheet_name or 'default'

        self._sheet = self._file.active
        self._sheet.title = self.sheet_name
        self.write(*self.header)

    def _import(self):
        import openpyxl
        import xlrd
        import os.path

        self.path = os.path
        self.openpyxl = openpyxl
        self.xlrd = xlrd


# TODO(dmvieira) not implemented
class Ods(PyHeaderSheet):
    """
    class that read ods files.
    Need reimplementing with new module. Peharps theses links:
    http://www.marco83.com/work/173/read-an-ods-file-with-python-and-odfpy/
    http://opendocumentfellowship.com/projects/odfpy
    """

    def __init__(self, name=None, header=list(), sheet_name=None):
        self.name = name
        self.header = header
        self.sheet_name = sheet_name
        super(Ods, self).__init__()

    def read_cell(self, x, y):
        raise NotImplementedError

    def write_cell(self, x, y, value, style=None):
        raise NotImplementedError

    def save(self, path=None):
        raise NotImplementedError

    def close(self, path=None):
        raise NotImplementedError

    def get_sheets(self):
        # self.sheets = [s.name for s in self._file.sheets]
        # return self.sheets
        return NotImplementedError

    def _open(self):
        # self._file = self.ezodf.opendoc(self.name)
        raise NotImplementedError

    def _import(self):
        # import ezodf
        # self.ezodf = ezodf
        raise NotImplementedError


class GSheet(PyHeaderSheet):
    """
    Class that read google spreadsheet files
    """

    def __init__(self, email, password, name=None, header=list(),
                 sheet_name=None, strip=False):
        """
        :param email: email
        :param password: password
        :param name: file name
        :param header: list with the header ['text1','text2','text3']
        :param sheet_name: sheet name
        :param strip: true to take spaces of values
        :return:
        """
        self.name = name
        self.email = email
        self.password = password
        self.header = header
        self.strip = strip
        self.sheet_name = sheet_name
        self._login()
        super(GSheet, self).__init__()

    def read_cell(self, x, y):
        """
        Reads the cell at position x+1 and y+1; return value
        :param x: line index
        :param y: coll index
        :return: {header: value}
        """
        if isinstance(self.header[y], tuple):
            header = self.header[y][0]
        else:
            header = self.header[y]
        x += 1
        y += 1
        if self.strip:
            self._sheet.cell(x, y).value = self._sheet.cell(x, y).value.strip()
        else:
            return {header: self._sheet.cell(x, y).value}

    def write_cell(self, x, y, value):
        """
        Writing value in the cell of x+1 and y+1 position
        :param x: line index
        :param y: coll index
        :param value: value to be written
        :return:
        """
        x += 1
        y += 1
        self._sheet.update_cell(x, y, value)

    def save(self, path=None):
        """

        :param path:
        :return:
        """
        if path:
            raise NotImplementedError
        return

    def _open(self):
        """
        Open the file; get sheets
        :return:
        """
        if not hasattr(self, '_file'):
            self._file = self.gc.open(self.name)
            self.sheet_names = self._file.worksheets()

    def _open_sheet(self):
        """
        Read the sheet, get value the header, get number columns and rows
        :return:
        """
        if self.sheet_name and not self.header:
            self._sheet = self._file.worksheet(self.sheet_name.title)
            self.ncols = self._sheet.col_count
            self.nrows = self._sheet.row_count
            for i in range(1, self.ncols+1):
                self.header = self.header + [self._sheet.cell(1, i).value]

    def _create(self):
        """

        :return:
        """
        raise NotImplementedError

    def _import(self):
        """
        Makes imports
        :return:
        """
        import os.path
        self.path = os.path

    def _login(self):
        """
        Login with your Google account
        :return:
        """
        import gspread
        self.gc = gspread.login(self.email, self.password)


def guess_type(filename,**kwargs):
    """ Utility function to call classes based on filename extension.
    Just usefull if you are reading the file and don't know file extension.
    You can pass kwargs and these args are passed to class only if they are
    used in class.
    """
    import os

    extension = os.path.splitext(filename)[1]
    case = {'.xls': Xls,
            '.xlsx': Xlsx,
            '.csv': Csv}
    if extension and case.get(extension.lower()):
        low_extension = extension.lower()
        new_kwargs = dict()
        class_name = case.get(low_extension)
        class_kwargs = class_name.__init__.func_code.co_names
        for kwarg in kwargs:
            if kwarg in class_kwargs:
                new_kwargs[kwarg] = kwargs[kwarg]
        return case.get(low_extension)(filename, **new_kwargs)
    else:
        raise Exception('No extension found')


def is_str_or_unicode(value):
    """
    Verifies if sting or unicode.
    :param value: value to be verified
    :return: True or None
    """
    if isinstance(value, str) or isinstance(value, unicode):
        return True

################################################################################
# run tests
################################################################################

if __name__ == '__main__':
    import doctest
    doctest.testmod()
